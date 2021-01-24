import openpyxl
import random
import time
import os
from log import init_logging
logger = init_logging(level='DEBUG')
time_now = time.strftime("%Y%m%d_%H_%M", time.localtime())
file_path="D:\python script\excel\data\ "+time_now
'''
if not os.path.exists(file_path):
    os.makedirs(file_path)
    print('文件夹创建完成  '+file_path)
'''
# 创建一个工作簿
wb = openpyxl.Workbook()
# 创建一个test_case的sheet表单s
wb.create_sheet('test_case1')
wb.create_sheet('test_case2')
# 保存为一个xlsx格式的文件
wb.save(file_path+"{}.xlsx".format('test_case'))#sava("/tmp/{}.xlsx".format(filename)),将路径添加进去就可以。

#将测试结果写入excel
def write_data(sheet_name,row,col,value):
    # 第一步：打开工作簿
    workbook1=openpyxl.load_workbook(file_path+'test_case.xlsx')
    sheet=workbook1[sheet_name]
    sheet.cell(row,col).value=value
    workbook1.save(file_path+"{}.xlsx".format('test_case'))

write_data('Sheet',1,1,'测试次数')
write_data('Sheet',1,2,'测试数据')
i=1
for i in range(1,10):
    i +=1
    data = random.randint(0,9)
    write_data('Sheet',i,1,i-1)
    write_data('Sheet',i,2,data)
logger.info('测试结束')
# 关闭工作薄
wb.close()
