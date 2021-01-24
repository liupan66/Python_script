import openpyxl

class Excel_write(object):
    def __init__(self,file_path,sheet_name):
        self.file_path=file_path
        self.sheet_name=sheet_name
        # 创建一个工作簿
        wb = openpyxl.Workbook()
        # 创建一个test_case的sheet表单s
        wb.create_sheet(self.sheet_name)
        # 保存为一个xlsx格式的文件
        wb.save(self.file_path+"{}.xlsx".format('test_case'))#sava("/tmp/{}.xlsx".format(filename)),将路径添加进去就可以。
    def write_data(self,row,col,value):
        # 第一步：打开工作簿
        workbook1=openpyxl.load_workbook(self.file_path+'test_case.xlsx')
        sheet=workbook1[self.sheet_name]
        sheet.cell(row,col).value=value
        workbook1.save(self.file_path+"{}.xlsx".format('test_case'))
    def close_excel(self):
        if hasattr(self, '_archive'):
            self._archive.close()

